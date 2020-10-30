from openpyxl import *


def editCell(file: object, Row: object, Column: object, text: object) -> object:
    wb = load_workbook(file)
    ws = wb["Booking"]
    testCell = ws.cell(Row, Column)
    testCell.value = text
    wb.save(file)


def checkCell(file: object, Row: object, Column: object, text: object) -> object:
    wb = load_workbook(file)
    ws = wb["Booking"]
    testCell = ws.cell(Row, Column)
    if testCell.value == text:
        return True
    else:
        return False


def getAllSeats(file):
    wb = load_workbook(file)
    ws = wb["Booking"]
    seats = []
    for row in ws.iter_rows(min_row=4, min_col=2, max_col=11, max_row=23):
        for cell in row:
            if cell.value is not None:
                seats.append([cell.row, cell.column])

    for row in ws.iter_rows(min_row=4, min_col=14, max_col=23, max_row=23):
        for cell in row:
            if cell.value is not None:
                seats.append([cell.row, cell.column])
    return seats
